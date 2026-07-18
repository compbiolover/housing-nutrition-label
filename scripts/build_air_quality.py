#!/usr/bin/env python3
"""Build the bundled air-quality tables (tract + county) for the Air Quality dimension.

Writes two artifacts so the runtime (``data/air_quality.py``) can resolve ambient
air quality at census-tract resolution (tract → county → unscored), with no
network call:

  • ``src/housing_label/data/air_quality_tracts.csv.gz`` — one row per census tract:
        geoid, pm25_ugm3, ozone_ppb
  • ``src/housing_label/data/air_quality.csv`` — one row per county (the fallback,
    and the sole carrier of the county-only radon layer):
        county_fips, pm25_ugm3, ozone_ppb, radon_zone

Sources — national, public-domain, joined on FIPS/GEOID:

  1. **PM2.5** — CDC Environmental Public Health Tracking downscaler-fused model.
     Tract: "Daily Census Tract-Level PM2.5" (Socrata ``vpk8-vfhm``, ``ds_pm_pred``).
     County: "Daily County-Level PM2.5" (``53mz-4zqd``, population-weighted
     ``pm25_pop_pred``). Annual mean for ``--year`` → µg/m³.
  2. **Ozone** — same downscaler basis. Tract: ``b72x-p96c`` (``ds_o3_pred``);
     county: ``3vxk-q2jk`` (``o3_pop_pred``). Annual mean of the daily max 8-hour → ppb.
  3. **Radon** — EPA "Map of Radon Zones" county classification (Zone 1 = highest
     predicted indoor level, ≥4 pCi/L; Zone 2 = 2–4; Zone 3 = <2). This is a
     *county-level* dataset with no finer public source, so it is bundled at the
     county grain and broadcast to a tract's county at runtime. EPA distributes it
     keyed by county *name + state*, joined to FIPS via the Census ``national_county2020``
     code list.

Default ``--year`` is 2021 — the latest year the tract PM2.5 series carries, so
tract and county share one vintage. All sources are fetched at build time
(dev-time only; the shipped artifacts are the CSVs). The tract/county pollutant
series are always pulled from the CDC Socrata API; the EPA radon workbook and the
Census county file can each be read from a local copy (``--radon-path`` /
``--census-path``) for a reproducible offline build.

Sources
-------
  CDC Tracking PM2.5 (tract / county):  vpk8-vfhm / 53mz-4zqd
  CDC Tracking Ozone (tract / county):  b72x-p96c / 3vxk-q2jk   (https://data.cdc.gov)
  EPA Map of Radon Zones:
    https://www.epa.gov/sites/default/files/2018-08/table_version_of_epa_radon_zones_by_county.xlsx
  Census county codes: https://www2.census.gov/geo/docs/reference/codes2020/national_county2020.txt

Run:  python scripts/build_air_quality.py            # fetch + write both tables
"""
from __future__ import annotations

import argparse
import csv
import gzip
import io
import pathlib
import re
import sys
import unicodedata

import requests

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

_DATA = pathlib.Path(__file__).resolve().parent.parent / "src" / "housing_label" / "data"
_OUT = _DATA / "air_quality.csv"
_TRACT_OUT = _DATA / "air_quality_tracts.csv.gz"

PM25_URL = "https://data.cdc.gov/resource/53mz-4zqd.json"
OZONE_URL = "https://data.cdc.gov/resource/3vxk-q2jk.json"
# Tract-level PM2.5 / ozone (same CDC downscaler model, census-tract grain).
PM25_TRACT_URL = "https://data.cdc.gov/resource/vpk8-vfhm.json"
OZONE_TRACT_URL = "https://data.cdc.gov/resource/b72x-p96c.json"
RADON_URL = ("https://www.epa.gov/sites/default/files/2018-08/"
             "table_version_of_epa_radon_zones_by_county.xlsx")
CENSUS_URL = "https://www2.census.gov/geo/docs/reference/codes2020/national_county2020.txt"

_HEADERS = {"User-Agent": "housing-nutrition-label/air-quality-build"}
_TIMEOUT = 120

# State USPS abbreviation → 2-digit state FIPS (radon join needs this; the EPA
# table is keyed by USPS abbr, the Census list by FIPS).
_STATE_ABBR_TO_FIPS = {
    "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06", "CO": "08",
    "CT": "09", "DE": "10", "DC": "11", "FL": "12", "GA": "13", "HI": "15",
    "ID": "16", "IL": "17", "IN": "18", "IA": "19", "KS": "20", "KY": "21",
    "LA": "22", "ME": "23", "MD": "24", "MA": "25", "MI": "26", "MN": "27",
    "MS": "28", "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
    "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38", "OH": "39",
    "OK": "40", "OR": "41", "PA": "42", "RI": "44", "SC": "45", "SD": "46",
    "TN": "47", "TX": "48", "UT": "49", "VT": "50", "VA": "51", "WA": "53",
    "WV": "54", "WI": "55", "WY": "56", "PR": "72",
}

# County-name suffixes Census appends but the EPA radon table omits. ("city" is
# deliberately NOT here — "Charles City"/"James City" are counties, not suffixes,
# and Census keeps "city" on independent cities too.) Longer phrases come first so
# a multi-word suffix ("city and borough") is stripped whole, never leaving
# "... city and" behind.
_SUFFIXES = re.compile(
    r"\s+(city and borough|census area|municipality|municipio|county|borough|"
    r"parish)$")

# EPA radon table quirks that no normalization can bridge: (state_fips, raw name)
# → the correct county name. The EPA workbook truncates a few leading letters
# and uses one legacy county name (Dade → Miami-Dade).
_RADON_ALIASES = {
    ("05", "an buren"): "van buren",   # AR Van Buren
    ("06", "entura"):   "ventura",     # CA Ventura
    ("37", "ance"):     "vance",       # NC Vance
    ("31", "hurston"):  "thurston",    # NE Thurston
    ("12", "dade"):     "miami dade",  # FL Miami-Dade (legacy name)
}


def _norm_name(name: str) -> str:
    """Normalize a county name for matching: strip accents, lowercase, drop the
    class suffix, fold punctuation (hyphens, periods, apostrophes) to spaces,
    collapse whitespace."""
    s = unicodedata.normalize("NFKD", name or "")
    s = "".join(c for c in s if not unicodedata.combining(c))  # Doña → Dona
    s = s.strip().lower()
    s = s.replace("st.", "saint").replace("ste.", "sainte")
    s = re.sub(r"[.\-'’]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = _SUFFIXES.sub("", s)
    return re.sub(r"\s+", " ", s).strip()


def _fips5(state_fips: str, county_fips: str) -> str:
    return f"{str(state_fips).strip().zfill(2)}{str(county_fips).strip().zfill(3)}"


def _socrata_county_annual(url: str, value_col: str, year: str) -> dict[str, float]:
    """Return {fips5: annual mean of `value_col`} for `year` (server-side avg)."""
    params = {
        "$select": f"statefips,countyfips,avg({value_col}) as v",
        "$where": f"year='{year}'",
        "$group": "statefips,countyfips",
        "$limit": "50000",
    }
    r = requests.get(url, params=params, headers=_HEADERS, timeout=_TIMEOUT)
    r.raise_for_status()
    out: dict[str, float] = {}
    for row in r.json():
        if row.get("v") is None:
            continue
        out[_fips5(row["statefips"], row["countyfips"])] = float(row["v"])
    return out


def _socrata_tract_annual(url: str, value_col: str, year: str) -> dict[str, float]:
    """Return {tract GEOID (11) → annual mean of `value_col`} for `year`.

    Server-side ``avg`` grouped by ``ctfips`` (~84k CONUS tracts) — one request,
    since Socrata returns the full grouped result under a high ``$limit``."""
    params = {
        "$select": f"ctfips,avg({value_col}) as v",
        "$where": f"year='{year}'",
        "$group": "ctfips",
        "$limit": "200000",
    }
    r = requests.get(url, params=params, headers=_HEADERS, timeout=_TIMEOUT)
    r.raise_for_status()
    out: dict[str, float] = {}
    for row in r.json():
        geoid = str(row.get("ctfips") or "").strip()
        if not geoid or row.get("v") is None:
            continue
        out[geoid.zfill(11)] = float(row["v"])
    return out


def _load_census_counties(path: str | None) -> dict[str, dict[str, str]]:
    """{state_fips: {normalized_county_name: fips5}} from the Census code list."""
    if path:
        text = pathlib.Path(path).read_text(encoding="utf-8")
    else:
        r = requests.get(CENSUS_URL, headers=_HEADERS, timeout=_TIMEOUT)
        r.raise_for_status()
        r.encoding = "utf-8"  # file is UTF-8; requests may guess latin-1 without a charset header
        text = r.text
    by_state: dict[str, dict[str, str]] = {}
    reader = csv.DictReader(io.StringIO(text), delimiter="|")
    for row in reader:
        sf = row["STATEFP"].strip()
        fips = _fips5(sf, row["COUNTYFP"])
        nm = _norm_name(row["COUNTYNAME"])
        d = by_state.setdefault(sf, {})
        d[nm] = fips                              # primary (keeps "city" for counties/ind. cities)
        d.setdefault(nm.replace(" ", ""), fips)   # De Kalb ↔ DeKalb, La Salle ↔ LaSalle
        if nm.endswith(" city"):
            # Independent city listed bare in the radon table ("Alexandria"). Use
            # setdefault so a same-named real county always keeps the bare key.
            d.setdefault(nm[:-5].strip(), fips)
    return by_state


def _match_fips(state_fips: str, norm: str, census: dict[str, dict[str, str]]) -> str | None:
    """Match a normalized county name to FIPS: exact, then space-insensitive."""
    idx = census.get(state_fips, {})
    return idx.get(norm) or idx.get(norm.replace(" ", ""))


def _load_radon(path: str | None,
                census: dict[str, dict[str, str]]) -> tuple[dict[str, int], list[str]]:
    """{fips5: radon_zone} from the EPA table, joined to FIPS via `census`.

    Returns (mapping, unmatched) where `unmatched` lists "State CountyName" the
    join could not resolve (logged, not fatal — those counties ship without a
    radon zone and are scored on PM2.5 + ozone alone at runtime, with radon's
    weight redistributed)."""
    if openpyxl is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read the EPA radon workbook.")
    if path:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    else:
        r = requests.get(RADON_URL, headers=_HEADERS, timeout=_TIMEOUT)
        r.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[str, int] = {}
    unmatched: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header: County Name | State | Radon Zone
            continue
        name, state, zone = (row + (None, None, None))[:3]
        if not name or not state or zone is None:
            continue
        sf = _STATE_ABBR_TO_FIPS.get(str(state).strip().upper())
        if not sf:
            continue
        norm = _norm_name(str(name))
        norm = _RADON_ALIASES.get((sf, norm), norm)
        fips = _match_fips(sf, norm, census)
        if fips is None:
            unmatched.append(f"{state} {name}")
            continue
        if fips in out:
            continue  # keep the first (correctly-labeled) row; skip stray duplicates
        try:
            out[fips] = int(zone)
        except (TypeError, ValueError):
            continue
    return out, unmatched


def _quantiles(vals: list[float], qs=(0.1, 0.25, 0.5, 0.75, 0.9, 0.95)) -> list[float]:
    s = sorted(vals)
    n = len(s)
    out = []
    for q in qs:
        idx = min(n - 1, int(round(q * (n - 1))))
        out.append(round(s[idx], 2))
    return out


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--year", default="2021",
                    help="modeled year (default 2021 — the latest the tract PM2.5 series carries)")
    ap.add_argument("--radon-path", help="local EPA radon .xlsx (skip fetch)")
    ap.add_argument("--census-path", help="local Census national_county file (skip fetch)")
    ap.add_argument("--out", default=str(_OUT), help="county CSV output path")
    ap.add_argument("--tracts-out", default=str(_TRACT_OUT), help="tract .csv.gz output path")
    args = ap.parse_args()

    # ── County layer (fallback + radon carrier) ───────────────────────────────
    print(f"PM2.5  county ← CDC Tracking 53mz-4zqd ({args.year}) …")
    pm25 = _socrata_county_annual(PM25_URL, "pm25_pop_pred", args.year)
    print(f"       {len(pm25)} counties")
    print(f"Ozone  county ← CDC Tracking 3vxk-q2jk ({args.year}) …")
    ozone = _socrata_county_annual(OZONE_URL, "o3_pop_pred", args.year)
    print(f"       {len(ozone)} counties")
    print("Radon  ← EPA Map of Radon Zones (+ Census county crosswalk) …")
    census = _load_census_counties(args.census_path)
    radon, unmatched = _load_radon(args.radon_path, census)
    print(f"       {len(radon)} counties matched"
          + (f", {len(unmatched)} unmatched" if unmatched else ""))
    if unmatched:
        print("       unmatched (ship without radon zone): "
              + ", ".join(sorted(unmatched)[:25])
              + (" …" if len(unmatched) > 25 else ""))

    fips_all = sorted(set(pm25) | set(ozone))
    with open(args.out, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county_fips", "pm25_ugm3", "ozone_ppb", "radon_zone"])
        for fips in fips_all:
            p, o, z = pm25.get(fips), ozone.get(fips), radon.get(fips)
            w.writerow([fips,
                        "" if p is None else f"{p:.2f}",
                        "" if o is None else f"{o:.1f}",
                        "" if z is None else z])
    print(f"Wrote {len(fips_all)} county rows → {args.out}")

    # ── Tract layer (primary resolution for PM2.5 + ozone) ────────────────────
    print(f"\nPM2.5  tract  ← CDC Tracking vpk8-vfhm ({args.year}) …")
    pm25_t = _socrata_tract_annual(PM25_TRACT_URL, "ds_pm_pred", args.year)
    print(f"       {len(pm25_t)} tracts")
    print(f"Ozone  tract  ← CDC Tracking b72x-p96c ({args.year}) …")
    ozone_t = _socrata_tract_annual(OZONE_TRACT_URL, "ds_o3_pred", args.year)
    print(f"       {len(ozone_t)} tracts")

    geoids = sorted(set(pm25_t) | set(ozone_t))
    with gzip.open(args.tracts_out, "wt", newline="") as f:
        w = csv.writer(f)
        w.writerow(["geoid", "pm25_ugm3", "ozone_ppb"])
        for g in geoids:
            p, o = pm25_t.get(g), ozone_t.get(g)
            w.writerow([g,
                        "" if p is None else f"{p:.2f}",
                        "" if o is None else f"{o:.1f}"])
    print(f"Wrote {len(geoids)} tract rows → {args.tracts_out}")

    # National quantiles that calibrate the scoring breakpoints in
    # data/air_quality.py. Anchor on the TRACT distribution (the primary
    # resolution) so a tract's score reads as a national percentile among tracts.
    if pm25_t:
        print(f"\nPM2.5 µg/m³ tract quantiles [10,25,50,75,90,95]: "
              f"{_quantiles(list(pm25_t.values()))}")
    if ozone_t:
        print(f"Ozone ppb   tract quantiles [10,25,50,75,90,95]: "
              f"{_quantiles(list(ozone_t.values()))}")
    if radon:
        from collections import Counter
        c = Counter(radon.values())
        print(f"Radon zone distribution: Z1={c.get(1,0)} Z2={c.get(2,0)} Z3={c.get(3,0)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
