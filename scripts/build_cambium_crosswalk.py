#!/usr/bin/env python3
"""Build the bundled county FIPS → Cambium GEA-region marginal grid factor table.

Writes ``src/housing_label/data/cambium_lrmer.csv`` (county_fips, gea_region,
lrmer_kgco2e_kwh) — the offline lookup that lets ``data/cambium.py`` return the
long-run **marginal** grid CO2e emission rate for any CONUS county, so the
environmental dimension can value solar/efficiency-avoided kWh at the marginal
rate (what actually gets turned off long-run) instead of the grid average.

Method (fully offline, government/lab-sourced)
----------------------------------------------
Read NREL's Cambium 2023 LRMER workbook (18 GEA regions, CONUS) and join its
per-region levelized marginal rate onto the workbook's own built-in county
crosswalk:

  1. ``Data - Annual`` tab, "Intermediate Levelization Calculations" block →
     each GEA region's levelized **Combustion CO2e** rate (kg CO2e / MWh of
     end-use demand), column F. Combustion (not combined) CO2e is used so the
     marginal rate is on the same stack-emissions basis as the eGRID *average*
     it is differenced against in the environmental model (eGRID total-output
     CO2e is likewise combustion-only) — see research/data-source-strengthening.
  2. ``County Mapping`` tab → State FIPS + County FIPS → Cambium GEA region.
  3. Join (2) onto (1), emit a 5-digit-FIPS → factor row per county.

The workbook is interactive: its levelized values reflect the user-input cells
on the ``Levelized LRMER`` tab. Those are validated against the target
configuration below (Mid-case / 2025 start / 20-yr levelization / 3% real
discount / 100-year AR6 GWP / end-use) and the build aborts on any mismatch, so
a re-saved workbook with different inputs can never silently change the bundled
factors.

Geographic scope: Cambium's GEA regions cover CONUS only. Alaska, Hawai'i, and
Puerto Rico counties are simply absent from the output and fall back (in
``data/cambium.py`` / the environmental model) to the eGRID average, i.e. no
marginal adjustment there.

Source (CC BY 4.0)
------------------
  • Gagnon, Pieter, et al. "Cambium 2023 Data." NREL, 2024.
    https://data.nrel.gov/submissions/230
    File: Cambium23_LRMER_GEAregions_0.xlsx
    Long-Run Marginal Emission Rates, CO2e, aggregated to the 18 GEA regions.

Cambium is reissued ~annually — treat these as dated constants and refresh
(re-run this script on the new workbook) on each Cambium release. The runtime
lookup signature is stable, so callers never change.

Run:  python scripts/build_cambium_crosswalk.py
      python scripts/build_cambium_crosswalk.py --xlsx local_Cambium23_LRMER_GEAregions.xlsx
"""

from __future__ import annotations

import argparse
import csv
import io
import pathlib
import sys

OUT = pathlib.Path(__file__).resolve().parents[1] / "src" / "housing_label" / "data" / "cambium_lrmer.csv"

CAMBIUM_URL = (
    "https://data.nrel.gov/system/files/230/"
    "1707947178-Cambium23_LRMER_GEAregions_0.xlsx"
)
HEADERS = {"User-Agent": "housing-nutrition-label/0.1 (cambium crosswalk build)"}

# The workbook is interactive; its levelized values are only the intended ones
# when its user-input cells hold this configuration. Verified before extraction.
EXPECTED_CONFIG = {
    "Scenario": "Mid-case",
    "Start year": 2025,
    "Evaluation period (years)": 20,
    "Discount rate (real)": 0.03,
    "Global Warming Potentials": "100-year (AR6)",
    "Location": "End-use",
}

KG_PER_MWH_TO_KG_PER_KWH = 1.0 / 1000.0


def _load_bytes(local: str | None) -> bytes:
    if local:
        return pathlib.Path(local).read_bytes()
    print(f"  downloading {CAMBIUM_URL}")
    import requests

    r = requests.get(CAMBIUM_URL, headers=HEADERS, timeout=300)
    r.raise_for_status()
    return r.content


def _validate_config(wb) -> None:
    """Abort unless the workbook's user-input cells match EXPECTED_CONFIG."""
    ws = wb["Levelized LRMER"]
    got = {ws.cell(r, 2).value: ws.cell(r, 4).value for r in range(7, 16)
           if ws.cell(r, 2).value}
    problems = []
    for key, want in EXPECTED_CONFIG.items():
        have = got.get(key)
        # Numeric compare with tolerance; string compare otherwise.
        if isinstance(want, (int, float)) and isinstance(have, (int, float)):
            ok = abs(float(have) - float(want)) < 1e-9
        else:
            ok = str(have).strip() == str(want)
        if not ok:
            problems.append(f"{key!r}: workbook={have!r} expected={want!r}")
    if problems:
        raise SystemExit(
            "Cambium workbook is not set to the target LRMER configuration:\n  "
            + "\n  ".join(problems)
            + "\nSet those inputs on the 'Levelized LRMER' tab, re-save, and re-run."
        )


def region_factors(wb) -> dict[str, float]:
    """GEA region → levelized Combustion CO2e marginal rate (kg CO2e / kWh).

    Reads the ``Data - Annual`` "Intermediate Levelization Calculations" block:
    column B is the region, column F the levelized Combustion CO2e in kg/MWh.
    """
    ws = wb["Data - Annual"]
    out: dict[str, float] = {}
    for r in range(6, 60):
        region = ws.cell(r, 2).value          # col B
        if region is None:
            break
        co2e_kg_mwh = ws.cell(r, 6).value     # col F — Combustion CO2e
        if co2e_kg_mwh is None:
            continue
        out[str(region).strip()] = float(co2e_kg_mwh) * KG_PER_MWH_TO_KG_PER_KWH
    return out


def county_regions(wb) -> dict[str, str]:
    """5-digit county FIPS → Cambium GEA region, from the ``County Mapping`` tab."""
    ws = wb["County Mapping"]
    out: dict[str, str] = {}
    rows = ws.iter_rows(min_row=2, max_col=7, values_only=True)
    for state_fips, county_fips, _county, _abbr, _state, _reeds, gea in rows:
        if gea is None or state_fips is None or county_fips is None:
            continue
        fips = f"{int(state_fips):02d}{int(county_fips):03d}"
        out[fips] = str(gea).strip()
    return out


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", help="Local Cambium23_LRMER_GEAregions xlsx (else download).")
    args = ap.parse_args()

    import openpyxl

    print("Loading Cambium 2023 LRMER workbook …")
    wb = openpyxl.load_workbook(io.BytesIO(_load_bytes(args.xlsx)),
                                read_only=True, data_only=True)

    print("Validating workbook configuration …")
    _validate_config(wb)

    factors = region_factors(wb)
    print(f"  {len(factors)} GEA regions "
          f"({min(factors.values()):.4f}–{max(factors.values()):.4f} kgCO2e/kWh)")

    crosswalk = county_regions(wb)
    print(f"  {len(crosswalk)} counties mapped")

    # Sanity: every mapped region must have a factor.
    orphans = sorted({g for g in crosswalk.values() if g not in factors})
    if orphans:
        raise SystemExit(f"County-mapping regions with no factor: {orphans}")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["county_fips", "gea_region", "lrmer_kgco2e_kwh"])
        for fips in sorted(crosswalk):
            gea = crosswalk[fips]
            w.writerow([fips, gea, round(factors[gea], 4)])
    print(f"Wrote {OUT} — {len(crosswalk)} counties.")

    if len(crosswalk) < 3000:
        print(f"WARNING: only {len(crosswalk)} counties resolved (expected ~3,100).",
              file=sys.stderr)


if __name__ == "__main__":
    main()
